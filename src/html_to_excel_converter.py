#!/usr/bin/env python3
"""
HTMLファイルをレイアウトを保持したままExcelファイルに変換するプログラム

このプログラムは以下の機能を提供します：
- HTMLファイルの読み込みと解析
- テーブルデータの抽出とExcelへの変換
- レイアウト情報の保持
- 複数のHTMLファイルの一括処理
- エラーハンドリングとログ機能
"""

import argparse
import logging
import os
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
import re

import pandas as pd
from bs4 import BeautifulSoup, Tag
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet


class HTMLToExcelConverter:
    """HTMLファイルをExcelファイルに変換するクラス"""
    
    def __init__(self, log_level: str = "INFO"):
        """
        初期化
        
        Args:
            log_level: ログレベル（DEBUG, INFO, WARNING, ERROR）
        """
        self.setup_logging(log_level)
        self.logger = logging.getLogger(__name__)
        
    def setup_logging(self, log_level: str) -> None:
        """ログ設定を初期化"""
        logging.basicConfig(
            level=getattr(logging, log_level.upper()),
            format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler(sys.stdout),
                logging.FileHandler('html_to_excel_conversion.log', encoding='utf-8')
            ]
        )
    
    def read_html_file(self, file_path: str) -> BeautifulSoup:
        """
        HTMLファイルを読み込んでBeautifulSoupオブジェクトを返す
        
        Args:
            file_path: HTMLファイルのパス
            
        Returns:
            BeautifulSoupオブジェクト
            
        Raises:
            FileNotFoundError: ファイルが見つからない場合
            Exception: ファイル読み込みエラー
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            soup = BeautifulSoup(content, 'html.parser')
            self.logger.info(f"HTMLファイルを読み込みました: {file_path}")
            return soup
        except FileNotFoundError:
            self.logger.error(f"ファイルが見つかりません: {file_path}")
            raise
        except Exception as e:
            self.logger.error(f"HTMLファイルの読み込みに失敗しました: {file_path}, エラー: {e}")
            raise
    
    def extract_tables(self, soup: BeautifulSoup) -> List[pd.DataFrame]:
        """
        HTMLからテーブルデータを抽出する
        
        Args:
            soup: BeautifulSoupオブジェクト
            
        Returns:
            テーブルデータのリスト（DataFrame）
        """
        tables = soup.find_all('table')
        dataframes = []
        
        for i, table in enumerate(tables):
            try:
                # テーブルをDataFrameに変換
                df = pd.read_html(str(table), encoding='utf-8')[0]
                dataframes.append(df)
                self.logger.info(f"テーブル {i+1} を抽出しました: {df.shape[0]}行 x {df.shape[1]}列")
            except Exception as e:
                self.logger.warning(f"テーブル {i+1} の抽出に失敗しました: {e}")
                continue
                
        return dataframes
    
    def extract_text_content(self, soup: BeautifulSoup) -> List[Dict[str, Any]]:
        """
        HTMLからテキストコンテンツを抽出して構造化データとして返す
        
        Args:
            soup: BeautifulSoupオブジェクト
            
        Returns:
            構造化されたテキストコンテンツのリスト
        """
        content = []
        
        # 見出しを抽出
        for level in range(1, 7):
            headers = soup.find_all(f'h{level}')
            for header in headers:
                content.append({
                    'type': f'heading_{level}',
                    'text': header.get_text().strip(),
                    'level': level,
                    'tag': f'h{level}'
                })
        
        # 段落を抽出
        paragraphs = soup.find_all('p')
        for p in paragraphs:
            text = p.get_text().strip()
            if text:
                content.append({
                    'type': 'paragraph',
                    'text': text,
                    'tag': 'p'
                })
        
        # リストを抽出
        lists = soup.find_all(['ul', 'ol'])
        for list_elem in lists:
            items = list_elem.find_all('li')
            list_items = [li.get_text().strip() for li in items if li.get_text().strip()]
            if list_items:
                content.append({
                    'type': 'list',
                    'items': list_items,
                    'tag': list_elem.name
                })
        
        # セクションを抽出
        sections = soup.find_all(['section', 'div'], class_=True)
        for section in sections:
            section_text = section.get_text().strip()
            if section_text and len(section_text) > 10:  # 短すぎるセクションは除外
                content.append({
                    'type': 'section',
                    'text': section_text,
                    'class': section.get('class', []),
                    'tag': section.name
                })
        
        self.logger.info(f"テキストコンテンツを抽出しました: {len(content)}個の要素")
        return content
    
    def create_excel_workbook(self, tables: List[pd.DataFrame], 
                            text_content: List[Dict[str, Any]], 
                            output_path: str) -> None:
        """
        Excelワークブックを作成して保存する
        
        Args:
            tables: テーブルデータのリスト
            text_content: テキストコンテンツのリスト
            output_path: 出力ファイルパス
        """
        wb = Workbook()
        
        # デフォルトシートを削除
        wb.remove(wb.active)
        
        # テーブル用のシートを作成
        if tables:
            for i, df in enumerate(tables):
                sheet_name = f"テーブル_{i+1}"
                ws = wb.create_sheet(title=sheet_name)
                self.add_table_to_worksheet(ws, df)
        
        # テキストコンテンツ用のシートを作成
        if text_content:
            ws = wb.create_sheet(title="コンテンツ")
            self.add_text_content_to_worksheet(ws, text_content)
        
        # サマリーシートを作成
        summary_ws = wb.create_sheet(title="サマリー", index=0)
        self.add_summary_to_worksheet(summary_ws, tables, text_content)
        
        # ファイルを保存
        wb.save(output_path)
        self.logger.info(f"Excelファイルを保存しました: {output_path}")
    
    def add_table_to_worksheet(self, ws: Worksheet, df: pd.DataFrame) -> None:
        """
        ワークシートにテーブルを追加する
        
        Args:
            ws: ワークシートオブジェクト
            df: データフレーム
        """
        # データをワークシートに追加
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # ヘッダーのスタイルを設定
        if ws.max_row > 0:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
        
        # 列幅を自動調整
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def add_text_content_to_worksheet(self, ws: Worksheet, 
                                    text_content: List[Dict[str, Any]]) -> None:
        """
        ワークシートにテキストコンテンツを追加する
        
        Args:
            ws: ワークシートオブジェクト
            text_content: テキストコンテンツのリスト
        """
        row = 1
        
        for content in text_content:
            if content['type'].startswith('heading'):
                # 見出し
                cell = ws.cell(row=row, column=1, value=content['text'])
                level = content.get('level', 1)
                font_size = max(16 - level, 10)
                cell.font = Font(bold=True, size=font_size, color="2F4F4F")
                cell.alignment = Alignment(horizontal="left", vertical="top")
                row += 1
                
            elif content['type'] == 'paragraph':
                # 段落
                cell = ws.cell(row=row, column=1, value=content['text'])
                cell.font = Font(size=11)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[row].height = max(15, len(content['text']) // 80 * 15)
                row += 1
                
            elif content['type'] == 'list':
                # リスト
                for item in content['items']:
                    cell = ws.cell(row=row, column=1, value=f"• {item}")
                    cell.font = Font(size=11)
                    cell.alignment = Alignment(horizontal="left", vertical="top", indent=1)
                    row += 1
                    
            elif content['type'] == 'section':
                # セクション
                cell = ws.cell(row=row, column=1, value=content['text'])
                cell.font = Font(size=11, italic=True)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                ws.row_dimensions[row].height = max(15, len(content['text']) // 80 * 15)
                row += 1
            
            row += 1  # 空行を追加
        
        # 列幅を設定
        ws.column_dimensions['A'].width = 100
    
    def add_summary_to_worksheet(self, ws: Worksheet, 
                               tables: List[pd.DataFrame], 
                               text_content: List[Dict[str, Any]]) -> None:
        """
        サマリーシートを作成する
        
        Args:
            ws: ワークシートオブジェクト
            tables: テーブルデータのリスト
            text_content: テキストコンテンツのリスト
        """
        # タイトル
        title_cell = ws.cell(row=1, column=1, value="HTML to Excel 変換サマリー")
        title_cell.font = Font(bold=True, size=16, color="2F4F4F")
        title_cell.alignment = Alignment(horizontal="center")
        
        # テーブル情報
        ws.cell(row=3, column=1, value="テーブル情報:").font = Font(bold=True)
        if tables:
            for i, df in enumerate(tables):
                ws.cell(row=4+i, column=1, value=f"テーブル {i+1}: {df.shape[0]}行 x {df.shape[1]}列")
        else:
            ws.cell(row=4, column=1, value="テーブルは見つかりませんでした")
        
        # コンテンツ情報
        content_row = 5 + len(tables) if tables else 5
        ws.cell(row=content_row, column=1, value="コンテンツ情報:").font = Font(bold=True)
        
        content_types = {}
        for content in text_content:
            content_type = content['type']
            content_types[content_type] = content_types.get(content_type, 0) + 1
        
        row = content_row + 1
        for content_type, count in content_types.items():
            ws.cell(row=row, column=1, value=f"{content_type}: {count}個")
            row += 1
        
        # 列幅を設定
        ws.column_dimensions['A'].width = 50
    
    def convert_html_to_excel(self, html_path: str, output_path: Optional[str] = None) -> str:
        """
        HTMLファイルをExcelファイルに変換する
        
        Args:
            html_path: 入力HTMLファイルのパス
            output_path: 出力Excelファイルのパス（省略時は自動生成）
            
        Returns:
            出力ファイルのパス
        """
        try:
            # 出力パスの生成
            if output_path is None:
                html_file = Path(html_path)
                output_path = html_file.with_suffix('.xlsx')
            
            # HTMLファイルを読み込み
            soup = self.read_html_file(html_path)
            
            # テーブルデータを抽出
            tables = self.extract_tables(soup)
            
            # テキストコンテンツを抽出
            text_content = self.extract_text_content(soup)
            
            # Excelファイルを作成
            self.create_excel_workbook(tables, text_content, output_path)
            
            self.logger.info(f"変換が完了しました: {html_path} -> {output_path}")
            return output_path
            
        except Exception as e:
            self.logger.error(f"変換に失敗しました: {html_path}, エラー: {e}")
            raise
    
    def batch_convert(self, input_dir: str, output_dir: Optional[str] = None) -> List[str]:
        """
        ディレクトリ内のHTMLファイルを一括変換する
        
        Args:
            input_dir: 入力ディレクトリのパス
            output_dir: 出力ディレクトリのパス（省略時は入力ディレクトリと同じ）
            
        Returns:
            変換されたファイルのパスのリスト
        """
        input_path = Path(input_dir)
        if output_dir is None:
            output_path = input_path
        else:
            output_path = Path(output_dir)
            output_path.mkdir(parents=True, exist_ok=True)
        
        html_files = list(input_path.glob("*.html"))
        converted_files = []
        
        self.logger.info(f"一括変換を開始します: {len(html_files)}個のHTMLファイル")
        
        for html_file in html_files:
            try:
                excel_file = output_path / f"{html_file.stem}.xlsx"
                self.convert_html_to_excel(str(html_file), str(excel_file))
                converted_files.append(str(excel_file))
            except Exception as e:
                self.logger.error(f"ファイルの変換に失敗しました: {html_file}, エラー: {e}")
                continue
        
        self.logger.info(f"一括変換が完了しました: {len(converted_files)}個のファイルを変換")
        return converted_files


def main():
    """メイン関数"""
    parser = argparse.ArgumentParser(
        description="HTMLファイルをレイアウトを保持したままExcelファイルに変換します"
    )
    parser.add_argument(
        "input", 
        help="入力HTMLファイルまたはディレクトリのパス"
    )
    parser.add_argument(
        "-o", "--output", 
        help="出力Excelファイルまたはディレクトリのパス"
    )
    parser.add_argument(
        "-l", "--log-level", 
        choices=["DEBUG", "INFO", "WARNING", "ERROR"], 
        default="INFO",
        help="ログレベル"
    )
    parser.add_argument(
        "--batch", 
        action="store_true",
        help="ディレクトリ内のHTMLファイルを一括変換"
    )
    
    args = parser.parse_args()
    
    # コンバーターを初期化
    converter = HTMLToExcelConverter(args.log_level)
    
    try:
        if args.batch:
            # 一括変換
            converted_files = converter.batch_convert(args.input, args.output)
            print(f"変換完了: {len(converted_files)}個のファイルを変換しました")
            for file_path in converted_files:
                print(f"  - {file_path}")
        else:
            # 単一ファイル変換
            output_file = converter.convert_html_to_excel(args.input, args.output)
            print(f"変換完了: {output_file}")
            
    except Exception as e:
        print(f"エラーが発生しました: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()


